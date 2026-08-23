#!/usr/bin/env python3
"""Convert docs/feature-docs.xlsx -> public/feature-data.json"""
import json, sys, os
try:
    import openpyxl
except ImportError:
    os.system("pip install openpyxl --break-system-packages -q")
    import openpyxl

ROOT = os.path.join(os.path.dirname(__file__), "..")
SRC  = os.path.join(ROOT, "docs", "feature-docs.xlsx")
DEST = os.path.join(ROOT, "public", "feature-data.json")

if not os.path.exists(SRC):
    print(f"[xlsx-to-json] {SRC} not found — skipping", file=sys.stderr)
    sys.exit(0)

wb = openpyxl.load_workbook(SRC, data_only=True)

# ── Feature Definitions ──────────────────────────────────────────────────────
def_sheet = next((wb[n] for n in wb.sheetnames if "Definition" in n), None)
definitions = []
if def_sheet:
    for row in def_sheet.iter_rows(min_row=4, values_only=True):
        fid = row[3]
        if not fid or not str(fid).startswith("FN-"):
            continue
        definitions.append({
            "ia":         str(row[0] or "").strip(),
            "d1":         str(row[1] or "").strip(),
            "d2":         str(row[2] or "").strip(),
            "id":         str(fid).strip(),
            "name":       str(row[4] or "").strip(),
            "target":     str(row[5] or "").strip(),
            "phase":      str(row[6] or "").strip(),
            "definition": str(row[7] or "").strip(),
            "revenue":    str(row[8] or "").strip(),
        })

# name lookup for resolving spec formulas
name_map = {d["id"]: d["name"] for d in definitions}

# ── Feature Specifications ───────────────────────────────────────────────────
spec_sheet = next((wb[n] for n in wb.sheetnames if "Spec" in n), None)
specs = []
if spec_sheet:
    for row in spec_sheet.iter_rows(min_row=2, values_only=True):
        fid = row[0]
        if not fid or not str(fid).startswith("FN-"):
            continue
        specs.append({
            "fid":       str(fid).strip(),
            "specId":    str(row[1] or "").strip(),
            "order":     int(row[2]) if row[2] is not None else 0,
            "name":      name_map.get(str(fid).strip(), ""),
            "type":      str(row[4] or "").strip(),
            "condition": str(row[5] or "").strip(),
            "process":   str(row[6] or "").strip(),
            "result":    str(row[7] or "").strip(),
            "note":      str(row[8] or "").strip(),
        })

out = {"definitions": definitions, "specs": specs}
with open(DEST, "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False, indent=2)

print(f"[xlsx-to-json] {len(definitions)} features, {len(specs)} specs → {DEST}")
